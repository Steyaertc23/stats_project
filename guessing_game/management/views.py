from os import path
from django.shortcuts import render, HttpResponse
from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth.models import User
from openpyxl import load_workbook
from django.conf import settings


def check_guessed_passwords(user:User):
    correct = 0
    for i, pw in enumerate(user.guesspasswordmodel_set.all()):
        guessed = user.guesspasswordmodel_set.get(pk=i+1)
        if pw.strip() == guessed.strip():
            correct += 1
    return correct

# Create your views here.

@user_passes_test(lambda user: user.is_superuser)
def generate_records(request):
    if len(User.objects.all()-2) == 60:
        wb = load_workbook(path.join(settings.BASE_DIR, "Password_guessing_game_results_stags.xlsx"))
        ws = wb['Results']
        ws["A1"] = "First Name"
        ws["B1"] = "Last Name"
        ws["C1"] = "Number of Correct passwords guessed"
        for i, user in enumerate(User.objects.all()[2::]):
            correct = check_guessed_passwords(user)
            

            ws[f"A{i+2}"] = user.first_name
            ws[f"B{i+2}"] = correct
            ws[f"C{i+2}"] = correct
        
        wb.save(path.join(settings.BASE_DIR, "Password_guessing_game_results_stags.xlsx"))
        return render(request, 'management/generate_records.html')
    #else:
    return HttpResponse("There Aren't 60 Responses Yet...")


def send_file(request):
    import os, mimetypes
    from wsgiref.util import FileWrapper
    from django.conf import settings
    filename     = os.path.join(settings.BASE_DIR, "Password_guessing_game_results_stags.xlsx")
    download_name = "Password_guessing_game_results_stags.xlsx"
    wrapper      = FileWrapper(open(filename, 'rb'))
    content_type = mimetypes.guess_type(filename)[0]
    response     = HttpResponse(wrapper,content_type=content_type)
    response['Content-Length']      = os.path.getsize(filename)    
    response['Content-Disposition'] = f"attachment; filename={download_name}"
    return response